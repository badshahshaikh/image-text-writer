import openpyxl
from PIL import Image, ImageDraw, ImageFont


path = "C:\\Users\\SR\\AppData\\Local\\Programs\\Python\\Python310\\code\\image_process\\Book1.xlsx"

wb=openpyxl.load_workbook(path)
sh1=wb['Sheet1']
row=sh1.max_row
column=sh1.max_column

i = 1

while i <= row:
  name = sh1.cell(i,1).value
  img = Image.open('image.png')
  d = ImageDraw.Draw(img)
  fnt = ImageFont.truetype("comic.ttf",25)
  d.text((250, 50), name, font=fnt, fill=(255, 255, 255) )
  img.save('final/' + name + '.png')
  print(name)
  i += 1
        
